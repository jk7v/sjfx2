import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import pandas as pd
import streamlit as st
import numpy as np
from datetime import datetime

from utils import dataframe_agent

# 页面性能优化配置
st.set_page_config(
    page_title="智能数据分析",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 设置页面样式
st.markdown("""
<style>
    /* 整体主题 */
    :root {
        --primary-color: #4A90E2;
        --secondary-color: #F5A623;
        --background-color: #F8F9FA;
        --text-color: #2C3E50;
        --border-color: #E9ECEF;
    }

    /* 整体背景和布局 */
    .stApp {
        background: var(--background-color);
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen-Sans, Ubuntu, Cantarell, 'Helvetica Neue', sans-serif;
    }
    
    /* 侧边栏优化 */
    .css-1d391kg, .css-1p05t8e {
        background-color: var(--background-color) !important;
        border-right: 1px solid var(--border-color);
    }
    
    /* 主要内容区域 */
    .block-container {
        padding: 1.5rem 2rem !important;
        max-width: 1200px;
        margin: 0 auto;
    }
    
    /* 卡片容器 */
    div[data-testid="stExpander"] {
        background: var(--background-color);
        border-radius: 4px;
        padding: 0.75rem;
        margin: 0.75rem 0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    }
    
    div[data-testid="stExpander"]:hover {
        background: #FFFFFF;
        box-shadow: 0 1px 4px rgba(0,0,0,0.03);
    }
    
    /* 输入控件美化 */
    .stSelectbox, .stTextInput>div>div {
        background: var(--background-color);
        border: 1px solid var(--border-color);
        border-radius: 4px;
    }
    
    .stSelectbox:hover, .stTextInput>div>div:hover {
        border-color: var(--primary-color);
        background: #FFFFFF;
    }
    
    /* 文件上传区域 */
    .stUploader {
        background: var(--background-color);
        border-radius: 4px;
        padding: 1rem;
        border: 1px dashed var(--border-color);
    }
    
    .stUploader:hover {
        border-color: var(--primary-color);
        background: #FFFFFF;
    }
    
    /* 标题和文字样式 */
    h1 {
        color: var(--text-color);
        font-size: 2rem;
        font-weight: 500;
        margin-bottom: 1rem;
    }
    
    h2, h3 {
        color: var(--text-color);
        font-weight: 400;
        margin: 0.75rem 0;
    }
    
    p {
        color: var(--text-color);
        line-height: 1.5;
    }
    
    /* 图表容器优化 */
    .stPlotlyChart {
        background: var(--background-color);
        border-radius: 4px;
        padding: 0.75rem;
        margin: 0.75rem 0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    }
    
    /* 按钮样式 */
    .stButton>button {
        background: var(--primary-color);
        color: white;
        border: none;
        border-radius: 4px;
        padding: 0.5rem 1rem;
        font-weight: 400;
    }
    
    .stButton>button:hover {
        background: #357ABD;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    
    /* 滚动条美化 */
    ::-webkit-scrollbar {
        width: 6px;
        height: 6px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--background-color);
    }
    
    ::-webkit-scrollbar-thumb {
        background: #CED4DA;
        border-radius: 3px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: #ADB5BD;
    }
</style>
""", unsafe_allow_html=True)

# 设置中文字体和样式
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False
sns.set_style("whitegrid")
sns.set_palette("husl")


@st.cache_data
def create_advanced_chart(input_data: dict, chart_type: str, title: str = "数据分析图表", x_label: str = "类别", y_label: str = "数值") -> None:
    """生成优化的统计图表"""
    columns = input_data["columns"]
    data = input_data["data"]
    
    # 设置自定义颜色方案
    color_sequence = px.colors.qualitative.Set3
    
    if chart_type == "bar":
        safe_columns = [str(col) for col in columns]
        safe_data = [float(val) if isinstance(val, (int, float)) else 0 for val in data]
        
        fig = go.Figure(data=[
            go.Bar(
                x=safe_columns,
                y=safe_data,
                marker=dict(
                    color=safe_data,
                    colorscale='Viridis',
                    showscale=True
                ),
                text=safe_data,
                texttemplate='%{text:.2f}',
                textposition='auto',
                hovertemplate='%{x}<br>%{y:.2f}'
            )
        ])
        
        fig.update_layout(
            title=dict(
                text=str(title),
                font=dict(size=24)
            ),
            xaxis_title=dict(
                text=str(x_label),
                font=dict(size=14)
            ),
            yaxis_title=dict(
                text=str(y_label),
                font=dict(size=14)
            ),
            height=500,
            showlegend=False,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white',
            paper_bgcolor='white',
            hoverlabel=dict(bgcolor='white')
        )
        
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': True})
        
    elif chart_type == "line":
        safe_columns = [str(col) for col in columns]
        safe_data = [float(val) if isinstance(val, (int, float)) else 0 for val in data]
        
        fig = go.Figure()
        
        # 添加主线
        fig.add_trace(
            go.Scatter(
                x=safe_columns,
                y=safe_data,
                mode='lines+markers',
                name='数据趋势',
                line=dict(width=3, color='rgb(66, 133, 244)'),
                marker=dict(
                    size=8,
                    color='rgb(66, 133, 244)',
                    symbol='circle'
                ),
                hovertemplate='%{x}<br>%{y:.2f}'
            )
        )
        
        # 添加范围区域
        fig.add_trace(
            go.Scatter(
                x=safe_columns,
                y=[y * 1.1 for y in safe_data],
                mode='lines',
                line=dict(width=0),
                showlegend=False,
                hoverinfo='skip'
            )
        )
        
        fig.add_trace(
            go.Scatter(
                x=safe_columns,
                y=[y * 0.9 for y in safe_data],
                mode='lines',
                line=dict(width=0),
                fillcolor='rgba(66, 133, 244, 0.2)',
                fill='tonexty',
                showlegend=False,
                hoverinfo='skip'
            )
        )
        
        fig.update_layout(
            title=dict(
                text=str(title),
                font=dict(size=24)
            ),
            xaxis_title=dict(
                text=str(x_label),
                font=dict(size=14)
            ),
            yaxis_title=dict(
                text=str(y_label),
                font=dict(size=14)
            ),
            height=500,
            showlegend=True,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white',
            paper_bgcolor='white',
            hoverlabel=dict(bgcolor='white'),
            legend=dict(
                yanchor="top",
                y=0.99,
                xanchor="left",
                x=0.01
            )
        )
        
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': True})
        
    elif chart_type == "pie":
        safe_columns = [str(col) for col in columns]
        safe_data = [float(val) if isinstance(val, (int, float)) else 0 for val in data]
        
        fig = go.Figure(data=[
            go.Pie(
                labels=safe_columns,
                values=safe_data,
                hole=0.4,
                textinfo='label+percent',
                textposition='outside',
                texttemplate='%{label}<br>%{percent:.1%}',
                marker=dict(colors=color_sequence),
                hovertemplate='%{label}<br>数值: %{value:.2f}<br>占比: %{percent:.1%}'
            )
        ])
        
        fig.update_layout(
            title=dict(
                text=str(title),
                font=dict(size=24)
            ),
            height=500,
            showlegend=True,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white',
            paper_bgcolor='white',
            hoverlabel=dict(bgcolor='white'),
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )
        
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': True})
        
    elif chart_type == "scatter":
        safe_columns = [str(col) for col in columns]
        safe_data = [float(val) if isinstance(val, (int, float)) else 0 for val in data]
        
        # 生成渐变色
        colors = [i/(len(safe_data)-1) for i in range(len(safe_data))]
        
        fig = go.Figure(data=[
            go.Scatter(
                x=list(range(len(safe_columns))),
                y=safe_data,
                mode='markers',
                marker=dict(
                    size=12,
                    color=colors,
                    colorscale='Viridis',
                    showscale=True,
                    line=dict(width=1, color='white')
                ),
                text=safe_columns,
                hovertemplate='%{text}<br>数值: %{y:.2f}'
            )
        ])
        
        fig.update_layout(
            title=dict(
                text=str(title),
                font=dict(size=24)
            ),
            xaxis=dict(
                title=dict(
                    text=str(x_label),
                    font=dict(size=14)
                ),
                tickvals=list(range(len(safe_columns))),
                ticktext=safe_columns
            ),
            yaxis_title=dict(
                text=str(y_label),
                font=dict(size=14)
            ),
            height=500,
            showlegend=False,
            margin=dict(t=50, b=50, l=50, r=50),
            plot_bgcolor='white',
            paper_bgcolor='white',
            hoverlabel=dict(bgcolor='white')
        )
        
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGray')
        
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': True})


@st.cache_data
def create_data_summary(df: "pd.DataFrame") -> None:
    """生成优化的数据摘要信息"""
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("数据行数", f"{len(df):,}")
    with col2:
        st.metric("数据列数", len(df.columns))
    with col3:
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        st.metric("数值列数", len(numeric_cols))
    with col4:
        missing_data = df.isnull().sum().sum()
        st.metric("缺失值", f"{missing_data:,}")


def create_correlation_heatmap(df: "pd.DataFrame") -> bool:
    """生成相关性热力图"""
    numeric_df = df.select_dtypes(include=[np.number])
    if len(numeric_df.columns) > 1:
        correlation_matrix = numeric_df.corr()
        
        fig = go.Figure(data=go.Heatmap(
            z=correlation_matrix.values,
            x=correlation_matrix.columns,
            y=correlation_matrix.columns,
            colorscale='RdBu',
            zmid=0,
            text=np.round(correlation_matrix.values, 2),
            texttemplate="%{text}",
            textfont={"size": 10},
            hoverongaps=False
        ))
        
        fig.update_layout(
            title=dict(
                text="数据相关性热力图",
                x=0.5,
                font=dict(size=18, family="Arial Black")
            ),
            height=500,
            margin=dict(t=80, b=80, l=80, r=80)
        )
        
        st.plotly_chart(fig, use_container_width=True)
        return True
    return False


# 页面配置

# 设置页面背景和样式
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    .stApp > header {
        background-color: transparent;
    }
    .stSidebar .sidebar-content {
        background: rgba(255, 255, 255, 0.8);
        backdrop-filter: blur(10px);
    }
    .stTextInput > div > div > input {
        background-color: rgba(255, 255, 255, 0.9);
    }
    .stSelectbox > div > div > div {
        background-color: rgba(255, 255, 255, 0.9);
    }
    .stFileUploader > div {
        background-color: rgba(255, 255, 255, 0.9);
    }
    .element-container {
        background-color: rgba(255, 255, 255, 0.7);
        border-radius: 10px;
        padding: 10px;\
        margin: 10px 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
</style>

<div style='text-align: center; padding: 30px; background: rgba(255, 255, 255, 0.9); border-radius: 15px; box-shadow: 0 8px 32px rgba(31, 38, 135, 0.15); backdrop-filter: blur(4px); -webkit-backdrop-filter: blur(4px); border: 1px solid rgba(255, 255, 255, 0.18); margin-bottom: 30px;'>
    <h1 style='background: linear-gradient(120deg, #1f77b4, #2ecc71); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-size: 3.5em; font-weight: 700; margin-bottom: 15px;'>🚀 数据分析智能体</h1>
    <p style='font-size: 1.3em; color: #555; margin-bottom: 20px; line-height: 1.6;'>智能化数据洞察 • 专业级可视化分析 • 一站式数据解决方案</p>
    <div style='width: 60px; height: 4px; background: linear-gradient(90deg, #1f77b4, #2ecc71); margin: 0 auto;'></div>
</div>
""", unsafe_allow_html=True)

# 侧边栏配置
with st.sidebar:
    st.markdown("### 📁 数据上传")
    option = st.radio("选择数据文件类型:", ("Excel", "CSV"), help="支持Excel和CSV格式的数据文件")
    file_type = "xlsx" if option == "Excel" else "csv"
    data = st.file_uploader(f"上传{option}数据文件", type=file_type)
    
    if data:
        st.success("✅ 文件上传成功！")
        
    st.markdown("### 🎨 图表配置")
    chart_title = st.text_input("图表标题", value="数据分析图表")
    x_axis_label = st.text_input("X轴标签", value="类别")
    y_axis_label = st.text_input("Y轴标签", value="数值")
    
    st.markdown("### ℹ️ 使用说明")
    st.info("""
    1. 上传Excel或CSV数据文件
    2. 查看数据概览和统计信息
    3. 输入分析问题或可视化需求
    4. 获得智能分析结果和图表
    """)

option = st.radio("请选择数据文件类型:", ("Excel", "CSV"), key="hidden", label_visibility="hidden")
file_type = "xlsx" if option == "Excel" else "csv"

if data:
    try:
        if file_type == "xlsx":
            wb = openpyxl.load_workbook(data)
            if len(wb.sheetnames) > 1:
                sheet_option = st.selectbox("选择要加载的工作表：", wb.sheetnames)
            else:
                sheet_option = wb.sheetnames[0]
            st.session_state["df"] = pd.read_excel(data, sheet_name=sheet_option)
        else:
            # 尝试不同编码方式读取CSV文件
            try:
                st.session_state["df"] = pd.read_csv(data, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    st.session_state["df"] = pd.read_csv(data, encoding='gbk')
                except UnicodeDecodeError:
                    st.session_state["df"] = pd.read_csv(data, encoding='latin-1')
        
        df = st.session_state["df"]
        
        # 数据概览
        st.markdown("## 📊 数据概览")
        create_data_summary(df)
        
        # 数据预览
        col1, col2 = st.columns([2, 1])
        
        with col1:
            with st.expander("📋 原始数据预览", expanded=True):
                st.dataframe(df, use_container_width=True, height=300)
        
        with col2:
            with st.expander("📈 数据统计信息", expanded=True):
                numeric_df = df.select_dtypes(include=[np.number])
                if not numeric_df.empty:
                    st.dataframe(numeric_df.describe(), use_container_width=True)
                else:
                    st.info("暂无数值型数据")
        
        # 相关性分析
        if create_correlation_heatmap(df):
            st.markdown("## 🔗 数据相关性分析")
        
        # 新增功能：数据分布分析
        st.markdown("## 📊 数据分布分析")
        col1, col2 = st.columns(2)
        with col1:
            numeric_columns = df.select_dtypes(include=[np.number]).columns
            if not numeric_columns.empty:
                selected_column = st.selectbox("选择要分析的数值列：", numeric_columns)
                fig = go.Figure()
                
                # 添加美化后的直方图
                fig.add_trace(go.Histogram(
                    x=df[selected_column],
                    nbinsx=30,
                    name="分布直方图",
                    marker=dict(
                        color='rgba(55, 128, 191, 0.7)',
                        line=dict(color='rgba(55, 128, 191, 1)', width=1)
                    ),
                    opacity=0.7,
                    hovertemplate='数值: %{x}<br>频数: %{y}<extra></extra>'
                ))
                
                # 添加美化后的小提琴图
                fig.add_trace(go.Violin(
                    x=df[selected_column],
                    name="密度分布",
                    side='positive',
                    line_color='rgba(231, 99, 250, 1)',
                    fillcolor='rgba(231, 99, 250, 0.5)',
                    points=False,
                    meanline=dict(visible=True, color='rgba(231, 99, 250, 1)'),
                    hovertemplate='数值: %{x}<br>密度: %{y}<extra></extra>'
                ))
                
                # 添加核密度估计曲线
                kde = df[selected_column].plot.kde()
                x_kde = kde.get_lines()[0].get_xdata()
                y_kde = kde.get_lines()[0].get_ydata()
                plt.close()
                
                fig.add_trace(go.Scatter(
                    x=x_kde,
                    y=y_kde,
                    name='核密度估计',
                    line=dict(color='rgba(50, 205, 50, 0.8)', width=2, dash='dot'),
                    hovertemplate='数值: %{x:.2f}<br>密度: %{y:.4f}<extra></extra>'
                ))
                
                # 更新布局
                fig.update_layout(
                    title=dict(
                        text=f"{selected_column}的分布情况",
                        font=dict(size=24, color='#333'),
                        x=0.5,
                        y=0.95
                    ),
                    showlegend=True,
                    height=500,
                    template='plotly_white',
                    margin=dict(t=100, b=50, l=50, r=50),
                    legend=dict(
                        yanchor="top",
                        y=0.99,
                        xanchor="right",
                        x=0.99,
                        bgcolor='rgba(255, 255, 255, 0.8)',
                        bordercolor='rgba(0, 0, 0, 0.2)',
                        borderwidth=1
                    ),
                    hoverlabel=dict(
                        bgcolor='white',
                        font_size=12,
                        font_family='Arial'
                    )
                )
                
                # 更新坐标轴
                fig.update_xaxes(
                    title=dict(text=selected_column, font=dict(size=14)),
                    showgrid=True,
                    gridwidth=1,
                    gridcolor='rgba(0, 0, 0, 0.1)',
                    zeroline=True,
                    zerolinewidth=1.5,
                    zerolinecolor='rgba(0, 0, 0, 0.3)'
                )
                fig.update_yaxes(
                    title=dict(text='频数/密度', font=dict(size=14)),
                    showgrid=True,
                    gridwidth=1,
                    gridcolor='rgba(0, 0, 0, 0.1)',
                    zeroline=True,
                    zerolinewidth=1.5,
                    zerolinecolor='rgba(0, 0, 0, 0.3)'
                )
                
                st.plotly_chart(fig, use_container_width=True, config={
                    'displayModeBar': True,
                    'displaylogo': False,
                    'modeBarButtonsToAdd': ['drawline', 'drawopenpath', 'eraseshape']
                })

        with col2:
            if not numeric_columns.empty:
                # 箱线图用于检测异常值
                fig = go.Figure()
                fig.add_trace(go.Box(
                    y=df[selected_column],
                    name="箱线图",
                    boxpoints='outliers',
                    marker_color='rgb(107, 174, 214)',
                    line_color='rgb(8, 81, 156)'
                ))
                fig.update_layout(
                    title_text=f"{selected_column}的异常值检测",
                    showlegend=True,
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)

        # 新增功能：时间序列分析
        st.markdown("## 📈 时间序列分析")
        date_columns = df.select_dtypes(include=['datetime64']).columns
        if not date_columns.empty:
            col1, col2 = st.columns(2)
            with col1:
                date_column = st.selectbox("选择时间列：", date_columns)
                numeric_column = st.selectbox("选择数值列：", numeric_columns)
                
                # 时间序列趋势图
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=df[date_column],
                    y=df[numeric_column],
                    mode='lines+markers',
                    name='实际值',
                    line=dict(color='rgb(31, 119, 180)')
                ))
                
                # 添加移动平均线
                ma_period = 7
                if len(df) >= ma_period:
                    ma = df[numeric_column].rolling(window=ma_period).mean()
                    fig.add_trace(go.Scatter(
                        x=df[date_column],
                        y=ma,
                        mode='lines',
                        name=f'{ma_period}日移动平均',
                        line=dict(color='rgb(255, 127, 14)', dash='dash')
                    ))
                
                fig.update_layout(
                    title=f"{numeric_column}的时间序列趋势",
                    xaxis_title="时间",
                    yaxis_title=numeric_column,
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # 季节性分析
                if len(df) >= 30:  # 确保有足够的数据进行季节性分析
                    df['月份'] = df[date_column].dt.month
                    monthly_avg = df.groupby('月份')[numeric_column].mean().reset_index()
                    
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(
                        x=monthly_avg['月份'],
                        y=monthly_avg[numeric_column],
                        mode='lines+markers',
                        name='月度平均',
                        line=dict(color='rgb(44, 160, 44)')
                    ))
                    
                    fig.update_layout(
                        title=f"{numeric_column}的季节性分析",
                        xaxis_title="月份",
                        yaxis_title=f"平均{numeric_column}",
                        height=400
                    )
                    st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("未检测到时间类型的列，无法进行时间序列分析。请确保您的数据包含日期时间列。")
        
    except Exception as e:
        st.error(f"❌ 数据加载失败: {str(e)}")
        st.stop()

# 智能分析区域
st.markdown("## 🤖 智能数据分析")

if "df" in st.session_state:
    
    st.markdown("### 💬 自定义分析")
    query = st.text_area(
        "请输入您的数据分析问题或可视化需求：",
        placeholder="例如：分析销售数据的月度趋势，或者制作产品类别的销量对比图表",
        height=100,
        disabled="df" not in st.session_state
    )
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        button = st.button("🚀 开始分析", use_container_width=True, type="primary")
else:
    st.info("👆 请先在侧边栏上传数据文件")
    button = False
    query = None

if button and not data:
    st.warning("⚠️ 请先上传数据文件")
    st.stop()

if query and button:
    with st.spinner("🤖 AI正在深度分析中，请稍等..."):
        try:
            result = dataframe_agent(st.session_state["df"], query)
            
            st.markdown("## 📋 分析结果")
            
            # 显示调试信息（如果存在）
            if "debug_info" in result:
                with st.expander("🔧 调试信息", expanded=False):
                    st.warning(result["debug_info"])
            
            if "error" in result:
                with st.expander("❌ 错误详情", expanded=False):
                    st.error(result["error"])
            
            if "answer" in result:
                st.markdown("### 💡 分析洞察")
                st.success(result["answer"])
            
            if "table" in result:
                st.markdown("### 📊 数据表格")
                try:
                    table_df = pd.DataFrame(result["table"]["data"], columns=result["table"]["columns"])
                    st.dataframe(table_df, use_container_width=True)
                    
                    # 提供下载选项
                    csv = table_df.to_csv(index=False, encoding='utf-8-sig')
                    st.download_button(
                        label="📥 下载表格数据",
                        data=csv,
                        file_name=f'analysis_result_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv',
                        mime='text/csv'
                    )
                except (KeyError, ValueError, pandas.errors.EmptyDataError) as table_err:
                    st.error(f"表格数据格式错误: {str(table_err)}")
                    st.json(result["table"])  # 显示原始数据结构
            
            if "bar" in result:
                st.markdown("### 📊 柱状图分析")
                try:
                    create_advanced_chart(result["bar"], "bar", chart_title, x_axis_label, y_axis_label)
                except Exception as chart_err:
                    st.error(f"柱状图生成错误: {str(chart_err)}")
                    st.json(result["bar"])  # 显示原始数据结构
            
            if "line" in result:
                st.markdown("### 📈 趋势线图")
                try:
                    create_advanced_chart(result["line"], "line", chart_title, x_axis_label, y_axis_label)
                except Exception as chart_err:
                    st.error(f"折线图生成错误: {str(chart_err)}")
                    st.json(result["line"])  # 显示原始数据结构
            
            if "pie" in result:
                st.markdown("### 🥧 饼图分析")
                try:
                    create_advanced_chart(result["pie"], "pie", chart_title, x_axis_label, y_axis_label)
                except Exception as chart_err:
                    st.error(f"饼图生成错误: {str(chart_err)}")
                    st.json(result["pie"])  # 显示原始数据结构
            
            if "scatter" in result:
                st.markdown("### 🔸 散点图分析")
                try:
                    create_advanced_chart(result["scatter"], "scatter", chart_title, x_axis_label, y_axis_label)
                except Exception as chart_err:
                    st.error(f"散点图生成错误: {str(chart_err)}")
                    st.json(result["scatter"])  # 显示原始数据结构
            
            # 显示完整的返回结果用于调试
            with st.expander("🔍 完整返回结果（调试用）", expanded=False):
                st.json(result)
                
        except Exception as e:
            st.error(f"❌ 分析过程中出现错误: {str(e)}")
            st.info("💡 请尝试重新表述您的问题或检查数据格式")
            # 显示详细错误信息
            with st.expander("错误详情", expanded=False):
                st.code(str(e))

# 页脚
st.markdown("""
<div style='text-align: center; color: #666; padding: 20px;'>
    <p>📈 数据分析智能体 让数据洞察更简单 | Powered by AI</p>
</div>
""", unsafe_allow_html=True)
st.markdown("### 🤖 AI 聊天")
if st.button("进入AI聊天页面", use_container_width=True):
    st.switch_page("pages/chat.py")
